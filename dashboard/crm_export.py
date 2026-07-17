from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SEGMENT_LABELS = {
    "all": "전체 분류",
    "new": "신규",
    "returning": "재방문",
    "potential_vip": "잠재 VIP",
    "vip": "VIP",
    "dormant": "휴면",
    "at_risk": "주의",
}
STATUS_LABELS = {"all": "전체 고객", "normal": "일반", "restricted": "예약 제한"}
SORT_LABELS = {
    "recent": "최근 방문순",
    "revenue": "매출 높은순",
    "visits": "방문 많은순",
    "no_show": "노쇼 높은순",
    "risk": "운영 위험 높은순",
    "name": "고객 이름순",
}

HEADERS = [
    "고객 ID", "고객명", "전화번호", "이메일", "고객 분류", "분류 사유",
    "완료 방문", "전체 예약", "누적 매출", "평균 객단가", "최근 방문",
    "다음 예약", "취소 횟수", "취소율", "노쇼 횟수", "노쇼율",
    "선호 서비스", "선호 디자이너", "예약 제한",
]


def build_customer_crm_workbook(customer_rows, *, query="", segment="all", status="all", sort_key="recent"):
    """Build a formatted CRM workbook from the exact rows shown on the admin screen."""
    wb = Workbook()
    ws = wb.active
    ws.title = "고객 현황"
    ws.sheet_view.showGridLines = False

    last_column = get_column_letter(len(HEADERS))
    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = "Salon De Nature 고객 CRM 현황"
    ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="174C3C")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_column}2")
    ws["A2"] = (
        f"검색어: {query or '-'}  |  고객 분류: {SEGMENT_LABELS[segment]}  |  "
        f"예약 제한: {STATUS_LABELS[status]}  |  정렬: {SORT_LABELS[sort_key]}"
    )
    ws["A2"].font = Font(name="맑은 고딕", size=10, color="475467")
    ws["A2"].fill = PatternFill("solid", fgColor="F3F6F4")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A3:{last_column}3")
    ws["A3"] = f"내보낸 고객 수: {len(customer_rows)}명  |  생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = Font(name="맑은 고딕", size=10, color="667085")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.append([])
    ws.append(HEADERS)
    header_row = 5
    data_start = header_row + 1

    thin_gray = Side(style="thin", color="D9E0DC")
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor="DDEBE5")
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="173F34")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)
    ws.row_dimensions[header_row].height = 28

    for row in customer_rows:
        customer = row["customer"]
        ws.append([
            customer.id,
            customer.name or "",
            customer.phone or "",
            customer.email or "",
            row["segment_label"],
            row["segment_reason"],
            row["completed_count"],
            row["booking_count"],
            row["total_revenue"],
            row["average_ticket"],
            row["last_visit_at"],
            row["next_booking_at"],
            row["cancelled_count"],
            row["cancellation_rate"] / 100,
            row["no_show_count"],
            row["no_show_rate"] / 100,
            row["preferred_service"] or "",
            row["preferred_staff"] or "",
            "Y" if customer.booking_restricted else "N",
        ])

    data_end = ws.max_row
    if data_end >= data_start:
        table = Table(displayName="CustomerCRMTable", ref=f"A{header_row}:{last_column}{data_end}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

        for row_cells in ws.iter_rows(min_row=data_start, max_row=data_end, min_col=1, max_col=len(HEADERS)):
            for cell in row_cells:
                cell.font = Font(name="맑은 고딕", size=10, color="1F2937")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=Side(style="hair", color="E8ECEA"))

        for row_number in range(data_start, data_end + 1):
            ws.cell(row_number, 9).number_format = '"₩"#,##0'
            ws.cell(row_number, 10).number_format = '"₩"#,##0'
            ws.cell(row_number, 11).number_format = "yyyy-mm-dd"
            ws.cell(row_number, 12).number_format = "yyyy-mm-dd hh:mm"
            ws.cell(row_number, 14).number_format = "0.0%"
            ws.cell(row_number, 16).number_format = "0.0%"

    widths = [10, 18, 16, 28, 14, 42, 12, 12, 15, 15, 14, 19, 11, 11, 11, 11, 22, 18, 11]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = f"A{header_row}:{last_column}{max(data_end, header_row)}"
    ws.print_title_rows = f"1:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
